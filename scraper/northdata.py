import openpyxl as op
import dryscrape
from bs4 import BeautifulSoup
import urllib as ul
dryscrape.start_xvfb()                                                         # Start dryscrape session
session = dryscrape.Session()

def northdata(input,col):
	wb = op.load_workbook(input)
	print type(wb)
	sheet = wb['Sheet1']

	# searchinput = session.at_xpath('//*[@id="search"]/div[1]/form/div/div[1]/input')
	# searchform = session.at_xpath('//*[@id="search"]/div[1]/form')
	for i in range(3,4):
		url = "https://www.northdata.de/"
		firma = sheet.cell(row=i, column=1).value
		firma = ul.quote(firma)
		searchurl = url + firma
		print searchurl
		session.visit(searchurl)
		session.render('north.png', width = 1900 , height = 1024)
		body = session.body()
		soup = BeautifulSoup(body)
		#print soup.prettify()
		for data in soup.find_all('div', class_='tab_content'):
			print data
		sheet.cell(row=i, column=col).value = 'Test'
	out = input + '_update.xlsx'
	wb.save(out)

northdata('karlsruhe.xlsx',10)
